"""Работа с Excel/CSV: чтение списков пользователей и формирование отчётов."""
from __future__ import annotations

import csv
import io
import re
from typing import Any, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook

_TOKEN_USERNAME = re.compile(r"(?:^|[^A-Za-z0-9_])@?([A-Za-z][A-Za-z0-9_]{4,31})$", re.IGNORECASE)
_TOKEN_TME = re.compile(r"(?:t(?:elegram)?\.(?:me|dog)/)(?:@?)([A-Za-z][A-Za-z0-9_]{4,31})", re.IGNORECASE)
_TOKEN_ID = re.compile(r"^\d{5,}$")
_TOKEN_PHONE = re.compile(r"^\+?\d{7,15}$")


def _tokens_from_cell(value: Any) -> List[str]:
    """Достаёт идентификаторы пользователей из одной ячейки."""
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    found: List[str] = []
    m = _TOKEN_TME.search(text)
    if m:
        found.append("@" + m.group(1))
        return found
    for part in re.split(r"[\s,;]+", text):
        part = part.strip()
        if not part:
            continue
        if _TOKEN_ID.match(part) or _TOKEN_PHONE.match(part):
            found.append(part.lstrip("+") if _TOKEN_PHONE.match(part) and not part.startswith("+") else part)
            continue
        m2 = _TOKEN_USERNAME.match(part)
        if m2 and part.startswith("@"):
            found.append("@" + m2.group(1))
            continue
        # «голое» слово не считаем юзернеймом — слишком много ложных срабатываний
    return found


def extract_users_from_file(filename: str, data: bytes) -> List[str]:
    """Читает xlsx/csv и возвращает список идентификаторов (без дублей, порядок сохранён)."""
    result: List[str] = []
    seen: set[str] = set()

    def add(token: str) -> None:
        key = token.lower().lstrip("@").lstrip("+")
        if key and key not in seen:
            seen.add(key)
            result.append(token if token.startswith("@") else token)

    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp1251"):
            try:
                text = data.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            return []
        for line in text.splitlines():
            for cell in re.split(r"[,;\t]", line):
                for token in _tokens_from_cell(cell):
                    add(token)
        return result

    # xlsx
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                for token in _tokens_from_cell(cell):
                    add(token)
    wb.close()
    return result


def extract_users_from_text(text: str) -> List[str]:
    """Достаёт идентификаторы из произвольного текста: @юзернеймы, t.me-ссылки, ID, телефоны."""
    result: List[str] = []
    seen: set[str] = set()

    def add(token: str) -> None:
        key = token.lower().lstrip("@")
        if key and key not in seen:
            seen.add(key)
            result.append(token)

    for chunk in re.split(r"[\n\r,;]+", text or ""):
        chunk = chunk.strip()
        if not chunk:
            continue
        m = _TOKEN_TME.search(chunk)
        if m:
            add("@" + m.group(1))
            continue
        for part in re.split(r"[\s\t]+", chunk):
            part = part.strip()
            if not part:
                continue
            if _TOKEN_ID.match(part) or _TOKEN_PHONE.match(part):
                add(part)
                continue
            if part.startswith("@") and re.match(r"^@[A-Za-z][A-Za-z0-9_]{4,31}$", part):
                add(part)
    return result


def build_users_xlsx(users: Iterable[dict]) -> bytes:
    """Excel-файл со списком пользователей (результат парсинга)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    headers = ["#", "ID", "Имя", "Юзернейм", "Последний визит", "Аватарка", "Источник"]
    ws.append(headers)
    for i, u in enumerate(users, 1):
        ws.append([
            i,
            u.get("id", ""),
            u.get("name", ""),
            u.get("username", ""),
            u.get("last_seen", ""),
            "да" if u.get("photo") else "нет",
            u.get("source", ""),
        ])
    widths = [6, 16, 28, 22, 24, 10, 26]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_report_xlsx(rows: List[Tuple[str, str, str]]) -> bytes:
    """Excel-отчёт: (идентификатор, статус, подробности)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    ws.append(["#", "Пользователь", "Статус", "Подробности"])
    for i, (ident, status, detail) in enumerate(rows, 1):
        ws.append([i, ident, status, detail])
    for idx, width in enumerate([6, 28, 34, 40], 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
